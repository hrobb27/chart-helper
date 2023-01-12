from datetime import date
from decouple import config
from lxml import html
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
import os
import re
import pickle
import requests
import spotipy
import argparse
import tqdm
import sys
import logging
from spotipy.oauth2 import SpotifyClientCredentials


albumLib = {} #IMPORTANT. Contains a KV pair of album ids (defined by PA) and album objects
chartLib = {}
class TqdmLoggingHandler(logging.StreamHandler):
    def __init__(self):
        logging.StreamHandler.__init__(self)

    def emit(self, record):
        msg = self.format(record)
        tqdm.tqdm.write(msg)
log = logging.getLogger(__name__)

log.setLevel(logging.DEBUG)
filehandler = logging.FileHandler('charthelper.log', 'w')
filehandler.setLevel(logging.DEBUG)
fileformatter = logging.Formatter('%(name)s | %(asctime)s | %(levelname)s | %(funcName)s | %(message)s', '%Y-%m-%d %H:%M:%S')
filehandler.setFormatter(fileformatter)
tqdmhandler = TqdmLoggingHandler()
formatter = logging.Formatter('%(levelname)s | %(funcName)s: %(message)s')
tqdmhandler.setFormatter(formatter)
log.addHandler(filehandler)
log.addHandler(tqdmhandler)
termsize = int(os.get_terminal_size()[0])
#Ranking: basically a class wrapper that contains the rank info and history for a single chart and album
#this class is used by the Album class to keep track of where an album appears in multiple charts
# (and how that's changed over time)
# Eg, Album will store unique instances of this class for "top 100 all time" and "top 100 prog metal"
# While there is a separate chart object, this is meant as a way to store information relative to
# a single album.  
class Ranking:
    def __init__(self, chart): #Set initial ranking state for some chart
        self.chart = chart
        self.ranking = -1
        self.lastUpdated = "" #date string of when this ranking was last made
        self.rankHistory = [] #log of previous rank numbers. Positive = rank. Neg = rejected status.
        self.isRejected = False  #def of rejection: when *this* album *was* on a chart once but not anymore
    def setRanking(self, rank, dateStr):  
        if len(self.rankHistory) < 1:
            log.debug("Ranking album for the first time")
            self.ranking = rank
            if rank == -1:
                self.isRejected = True
            self.lastUpdated = dateStr
            histEntry = [self.ranking, self.lastUpdated]
            log.debug("Added hist entry {s}".format(s=str(histEntry)))
            self.rankHistory.append(histEntry)
            
        else:
            log.debug("Adding new ranking")
            if rank == -1:  #Calling setRanking with a negative rank rejects album from chart
                self.isRejected = True
            else:
                self.isRejected = False
            histEntry = [self.ranking, self.lastUpdated]
            log.debug("Added hist entry {s}".format(s=str(histEntry)))
            self.rankHistory.append(histEntry)
            self.ranking = rank
            self.lastUpdated = dateStr

#Links: this class is also tied to the album class. While less necessary (you can view this as an 
# extension of the Album class anyhow), it's neater to put the bulk of the API calls in its own container
#A Links instance contains all the web links related to this album, as well as the library status
#it also generates some of these links. In the future I would like this to fetch RYM links,
#but they do not have a functional API, and scraping otherwise appears convoluted
class Links:
    def __init__(self, album, artist, year):
        self.album = album #just passing this for query purposes
        self.artist = artist
        self.year = year
        self.albumLink = "" 
        self.artistLink = ""
        self.spotifyLink = ""
        self.inLibrary = False
        self.onSpotify = False
        self.spotifyLink = ""
    def setSpotifyLink(self, spotifyLink): #manual streaming entry helper
        self.spotifyLink = spotifyLink
        self.onSpotify = True
    def notOnSpotify(self):  #manual streaming removal for an album
        self.spotifyLink = ""
        self.onSpotify = False
    def findInLibrary(self): #determine whether the album is in a user's music directory
        log.info("Seeking album in library")
        with open("settings.ini", 'r') as file:
            lines=file.readlines()
            libdir=lines[3].split('=')[1].strip('\n')
        root = libdir + self.artist[0] + '/' #This is fixed for now
        if os.path.isdir(root + self.artist):
            log.debug("Artist {a} found on disc".format(a=self.artist))
            path = root + self.artist + '/' + self.album
            if os.path.isdir(path):
                log.debug("{a} by {ar} found at {p}".format(a=self.album,ar=self.artist,p=path))
                self.inLibrary = True
                return
        log.info("Album not found, trying again with articles removed")
        doublecheck = self.artist.removeprefix("The ")
        if os.path.isdir(root + doublecheck):
            log.debug("Artist {a} found on disc".format(a=doublecheck))
            path = root + doublecheck + '/' + self.album
            if os.path.isdir(path):
                log.debug("{a} by {ar} found at {p}".format(a=self.album,ar=self.artist,p=path))
                self.inLibrary = True
                return
        log.info("Album not found on disc")
        self.inLibrary = False
    def findSpotify(self, spObj):
        log.info("Seeking spotify status for {a} by {ar}".format(a=self.album, ar=self.artist))
        
        artist = " ".join(re.split('\W+', self.artist))
        if not artist:
            log.debug("Artist sanitization failed, falling back")
            #WHY DID YOU PICK THAT BAND NAME???
            artist = self.artist
        else:
            artist= artist.group(0)
        log.debug("Sanitized artist name: {ar}".format(ar=artist))
        album = " ".join(re.split('\W+', self.album))
        if not album:
            log.debug("Album sanitization failed, falling back")
            album = self.album
        else:
            album = album.group(0)
        log.debug("Sanitized album name: {a}".format(a=album))
        albumQuery1 = "artist:{artist} album:{album}".format(artist=self.artist, album=self.album)
        log.info("Sending query for {s}".format(s=albumQuery1))
        res = spObj.search(albumQuery1, limit=3, type='album', market='US')
        link = ''
        #3 passes, querying spotify for the top 3 albums. This is pedantic
        if res['albums']['total'] == 0:
            log.info("No results for prev. query, trying again")
            albumQuery2 = "artist:{artist} year:{year}".format(artist=self.artist, year=self.year)
            log.info("Sending query for {s}".format(s=albumQuery2))
            res2 = spObj.search(albumQuery2, limit=3, type='album', market='US')
            if res2['albums']['total'] == 0:
                log.info("No results for prev. query, trying again (x2)")
                albumQuery3 = "album:{album} year:{year}".format(album=self.album, year=self.year)
                log.info("Sending query for {s}".format(s=albumQuery3))
                res3 = spObj.search(albumQuery3, limit=3, type='album', market='US')
                if res3['albums']['total'] == 0:
                    log.info("Could not find spotify link, setting status to false")
                    self.notOnSpotify()
                else:
                    log.info("Found spotify link!")
                    link = res3['albums']['items'][0]['external_urls']['spotify']
                    log.debug(link)
                    self.setSpotifyLink(link) 
            else:
                log.info("Found spotify link!")
                link = res2['albums']['items'][0]['external_urls']['spotify']
                log.debug(link)
                self.setSpotifyLink(link) 
        else:
            log.info("Found spotify link!")
            link = res['albums']['items'][0]['external_urls']['spotify']
            log.debug(link)
            self.setSpotifyLink(link) 
#Album Class
#Represents all the metadata for an album in a general package
#Invokes and handles 2 subclasses for rankings and links
#to do: move the other classes to be inner classes of album just to make things clearer
class Album:
    def __init__(self, idNo, title, artist, year ):
        self.idNo = idNo #PA id, retrieved from the PA link
        self.title = title
        self.artist = artist
        self.year = year
        self.locations = []
        self.duration = "Unknown"
        self.country = "Unknown"
        self.rank = {}
        self.links = Links(title, artist, year)
        self.rating = 0
        self.noRatings = 0
        self.houseListened = False
        self.bought = '?' #must be manually done
        self.numListeners = 0 
        self.profiles = {'Listened?' : '?'}

    def setListeningInfo(self, infoBlock, profileNames):#
        log.debug("Setting listening info")
        log.debug("Profiles: {pf}".format(pf=str(profileNames)))
        log.debug("Info: {pi}".format(pi=str(infoBlock)))
        if len(profileNames) == 0 and len(infoBlock) == 1:
            log.debug("Empty profile given, filling out generic profile instead")
            self.profiles['Listened?'] = infoBlock
            self.numListeners = len(infoBlock)
        elif len(profileNames) == len(infoBlock):
            log.debug("Filling out profile info")
            for i in range(len(profileNames)):
                self.profiles[profileNames[i]] = infoBlock[i] #local storage
            self.numListeners = len(infoBlock)
    def addRanking(self, chartName, rank, dateStr):#assign a ranking to a chart (str) and rank (positive int if on chart, negative for reject)
        log.debug("Adding rank {r} to album {t} on chart {c}".format(r=rank,t=self.title,c=chartName))
        if chartName not in self.rank.keys():
            log.debug("Album is new to chart")
            ranking = Ranking(chartName)
            ranking.setRanking(rank, dateStr)
            self.rank[chartName] = ranking
        else:
            log.debug("Updating rank for album")
            self.rank[chartName].setRanking(rank, dateStr)
    def deleteRanking(self, chartName): #only use when deleting a chart altogether
        return self.rank.pop(chartName, None)
    def enterRanking(self, rankObj):
        self.rank[rankObj.chart] = rankObj
    def syncListened(self, newBlock, newListeners, literalRule=False):
        log.debug("Syncing listening info")
        if len(newBlock) == len(newListeners):
            for i in range(len(newBlock)):
                listener = newListeners[i]
                if listener in self.profiles.keys():
                    log.debug("Updating existing profile entry for {l}".format(l=listener))
                    if literalRule:
                        old = self.profiles[listener]
                        self.profiles[listener] = newBlock[i]
                        log.debug("Overwrote {old} with {new}".format(old=old, new=newBlock[i]))
                    elif newBlock[i] == "Yes":
                        log.debug("Updated profile entry to Yes")
                        self.profiles[listener] = newBlock[i]
                    elif newBlock[i] == "No" and self.profiles[listener] == "?":
                        self.profiles[listener] = newBlock[i]
                    else:
                        log.debug("No change")
                else:
                    log.debug("New listener profile detected: {n}".format(n=listener))
                    self.profiles[listener] = newBlock[i]
def addTimes(t1, t2):
    log.debug("Adding times {t1} and {t2}".format(t1=t1, t2=t2))
    first = t1.split(':')
    second = t2.split(':')
    mins = int(first[0]) + int(second[0])
    secs = int(first[1]) + int(second[1])
    if secs > 59:
        mins += secs // 60
        secs = secs % 60
    if secs < 10:
        secs = '0' + str(secs)
    res = str(mins) + ":" + str(secs)
    log.debug("Result: {res}".format(res=res))
    return res

class Chart:
    def __init__(self, name, size, link, date):
        self.scanned = False
        self.name = name
        self.link = link
        self.isRejectChart = False
        if link == "REJECT":
            self.isRejectChart = True
        self.date = date
        self.size = size 
        self.entries = [None] * size
        self.heading = ['Artist', 'Album', 'Year', 'PA Genre', 'PA Rating', '# of Ratings', 'Album Length', 'Country', 'On Spotify?', 'Bought?', 'In Library?', 'House Listened?']
    def resize(self, newSize):
        log.debug("Resizing {name} to {size}".format(name=self.name, size=newSize))
        oldEntries = self.entries.copy()
        oldLen = len(self.entries)
        if newSize >= oldLen:
            tail = [None] * (newSize - oldLen)
            if not (len(tail) == 0):
                self.entries.append(tail)
            log.debug("Expanded entries from {old} to {new}".format(old=oldLen, new=newSize))
        else:
            end = newSize - oldLen
            self.entries = self.entries[:end]
            log.debug("Truncated entries from {old} to {new}".format(old=oldLen, new=newSize))
        self.size = newSize
        return oldEntries
    def addAlbum(self, albumID, rank):
        self.entries[rank - 1] = albumID#rank starts at 1, so we adjust to get the proper 0 indexing
        albumLib[albumID].addRanking(self.name, rank, self.date)
    def writeChart(self, sheet, profileIDs = ["Listened?"]):
        log.info("Writing chart {name} to sheet {sname} with profiles {pf}".format(name=self.name, sname=sheet.title, pf=str(profileIDs)))
        headerFont = Font('Times New Roman', 11, True, underline='single')
        heading = ['Artist', 'Album', 'Year', 'PA Genre', 'PA Rating', '# of Ratings', 'Album Length', 'Country'] + profileIDs + [ 'On Spotify?', 'Bought?', 'In Library?', 'House Listened?']
        log.debug(str(heading))
        for i in range(len(heading)):
            addr = str(get_column_letter(i + 1))+'1'
            sheet[addr] = heading[i]
            sheet[addr].font = headerFont
        plainFont = Font('Times New Roman', 11)
        if self.isRejectChart:
            log.info("Chart is a reject chart")
            rowCount = 2
            readCount = 0
            for rejSection in tqdm.tqdm(self.rejSections, total=len(self.rejSections), unit="Reject Section"):
                date = rejSection[0]
                albumCount = rejSection[1]
                sheet['A'+str(rowCount)] = date
                rowCount += 1
                with tqdm.tqdm(total=albumCount, position=1, bar_format='{desc}', desc='writeChart: Writing section {d}'.format(d=date)) as desc:
                    for i in tqdm.tqdm(range(albumCount), total=albumCount, unit="album", position=0):
                        a = albumLib[self.entries[readCount + i]]
                        r = str(rowCount)
                        sheet['A'+r].font = plainFont
                        sheet['A'+r]  = a.artist
                        sheet['A'+r].hyperlink = a.links.artistLink
                        sheet['B'+r].font = plainFont
                        sheet['B'+r]  = a.title
                        sheet['B'+r].hyperlink = a.links.albumLink
                        sheet['C'+r].font = plainFont
                        sheet['C'+r]  = a.year
                        sheet['D'+r].font = plainFont
                        sheet['D'+r]  = a.genre
                        sheet['E'+r].font = plainFont
                        sheet['E'+r]  = a.rating
                        sheet['F'+r].font = plainFont
                        sheet['F'+r]  = a.noRatings
                        sheet['G'+r].font = plainFont
                        sheet['G'+r]  = a.duration
                        sheet['H'+r].font = plainFont
                        sheet['H'+r]  = a.country
                        sheet['I'+r].font = plainFont
                        desc.set_description("writeChart: Section {d}: writing {a} by {ar} onto row {r}".format(d=date, a=a.title, ar=a.artist, r=r))
                        for j in range(len(profileIDs)):
                            listener_col = get_column_letter(9 + j) + r
                            thisAlbum = albumLib[a.idNo]
                            thisProfile = profileIDs[j]
                            if thisProfile in thisAlbum.profiles.keys():
                                sheet[listener_col] = thisAlbum.profiles[thisProfile] #yes, no, or ?
                            else:
                                sheet[listener_col] = "?"
                            sheet[listener_col].font = plainFont
                        offset = 9 + len(profileIDs)
                        next_col = get_column_letter(offset)
                        sheet[next_col+r].font = plainFont
                        if a.links.onSpotify:
                            sheet[next_col+r] = 'Yes'
                            sheet[next_col+r].hyperlink = a.links.spotifyLink
                            sheet[next_col+r].style='Hyperlink'
                        else:
                            sheet[next_col+r] = 'No'
                        sheet[next_col+r].font = plainFont
                        next_col = get_column_letter(offset + 1)
                        sheet[next_col+r].font = plainFont
                        sheet[next_col+r] = 'Yes' if a.bought == "Yes" else 'No'
                        next_col = get_column_letter(offset + 2)
                        sheet[next_col+r].font = plainFont
                        sheet[next_col+r] = 'Yes' if a.links.inLibrary == "Yes" else 'No'
                        next_col = get_column_letter(offset + 3)
                        sheet[next_col +r].font = plainFont
                        sheet[next_col +r] = 'Yes' if a.houseListened == "Yes" else 'No'
                        rowCount += 1
                    desc.set_description("writeChart: Completed writing chart {name} with {noAlbums} albums".format(name=self.name, noAlbums=self.size))
            return 
        else:
            with tqdm.tqdm(total=len(self.entries), position=1, bar_format='{desc}', desc='writeChart: Beginning process') as desc:
                for row in tqdm.tqdm(range(len(self.entries))):
                    a = albumLib[self.entries[row]]
                    r = str(row+2)
                    sheet['A'+r].font = plainFont
                    sheet['A'+r]  = a.artist
                    sheet['A'+r].hyperlink = a.links.artistLink
                    sheet['B'+r].font = plainFont
                    sheet['B'+r]  = a.title
                    sheet['B'+r].hyperlink = a.links.albumLink
                    sheet['C'+r].font = plainFont
                    sheet['C'+r]  = a.year
                    sheet['D'+r].font = plainFont
                    sheet['D'+r]  = a.genre
                    sheet['E'+r].font = plainFont
                    sheet['E'+r]  = a.rating
                    sheet['F'+r].font = plainFont
                    sheet['F'+r]  = a.noRatings
                    sheet['G'+r].font = plainFont
                    sheet['G'+r]  = a.duration
                    sheet['H'+r].font = plainFont
                    sheet['H'+r]  = a.country
                    sheet['I'+r].font = plainFont
                    desc.set_description("writeChart: Writing {a} by {ar} to row {r}".format(a=a.title,ar=a.artist,r=r))
                    for i in range(len(profileIDs)):
                        listener_col = get_column_letter(9 + i) + r
                        thisAlbum = albumLib[a.idNo]
                        thisProfile = profileIDs[i]
                        if thisProfile in thisAlbum.profiles.keys():
                            sheet[listener_col] = thisAlbum.profiles[thisProfile] #yes, no, or ?
                        else:
                            sheet[listener_col] = "?"
                        sheet[listener_col].font = plainFont
                    offset = 9 + len(profileIDs)
                    next_col = get_column_letter(offset)
                    sheet[next_col+r].font = plainFont
                    if a.links.onSpotify:
                        sheet[next_col+r] = 'Yes'
                        sheet[next_col+r].hyperlink = a.links.spotifyLink
                        sheet[next_col+r].style='Hyperlink'
                    else:
                        sheet[next_col+r] = 'No'
                    sheet[next_col+r].font = plainFont
                    next_col = get_column_letter(offset + 1)
                    sheet[next_col+r].font = plainFont
                    sheet[next_col+r] = 'Yes' if a.bought == "Yes" else 'No'
                    next_col = get_column_letter(offset + 2)
                    sheet[next_col+r].font = plainFont
                    sheet[next_col+r] = 'Yes' if a.links.inLibrary == "Yes" else 'No'
                    next_col = get_column_letter(offset + 3)
                    sheet[next_col +r].font = plainFont
                    sheet[next_col +r] = 'Yes' if a.houseListened == "Yes" else 'No'
                desc.set_description("writeChart: Completed writing chart {name} with {noAlbums} albums".format(name=self.name, noAlbums=self.size))
        return 

def getRejectChart(chartname):
    global date #I actually have no idea why it doesn't have date inhereited
    log.info("Creating reject chart for {name}".format(name=chartname))
    dateStr = date.today().strftime("w%m/%d/%Y")
    rejChartName = chartname + " REJECT"
    log.info("New reject chart name: {name}".format(name=rejChartName))
    if chartname in chartLib.keys():
        dates = {}
        rejCount = 0
        with tqdm.tqdm(total=len(albumLib.values()), position=1, bar_format='{desc}', desc='getRejectChart: Beginning process') as desc:
            for album in tqdm.tqdm(albumLib.values(), position=0, unit="album", total=len(albumLib.values())): #brute force I'm afraid
                if chartname in album.rank.keys():
                    log.debug("Found album {a} with history in {c}".format(a=album.title, c=chartname))
                    desc.set_description("getRejectChart: found album {a} with history in {c}".format(a=album.title, c=chartname))
                    thisRank = album.rank[chartname]
                    if thisRank.ranking == -1:
                        desc.set_description("getRejectChart: album {a} was rejected from {c}".format(a=album.title, c=chartname))
                        rejCount += 1
                        hist = reversed(thisRank.rankHistory)
                        prev = thisRank.lastUpdated
                        for entry in hist: #find the start of album's reject period
                            if entry[0] == -1: 
                                prev = entry[1] 
                            else:
                                break
                        desc.set_description("getRejectChart: album {a} was first rejected on {d}".format(a=album.title, d=prev))
                        if prev not in dates.keys():
                            log.info("Adding new reject section for date {d}".format(d=prev))
                            dates[prev] = [album]
                        else:
                            log.info("Appending album to date {d}".format(d=prev))
                            dates[prev].append(album)
        log.info("Finished finding rejects")
        datelist = sorted(dates.keys(), key= lambda dstr : dstr[6:10]+'/'+dstr[3:5]+'/'+dstr[0:2])
        #Sorts dates in oneliner fashion
        i = 1
        rejSections = []
        for date in datelist:
            noRejs = len(dates[date])
            log.debug("Rejected {n} albums from {c} on {d}".format(n=noRejs, c=chartname, d=date))
            rejSections.append((date, noRejs))
        rejChart = Chart(rejChartName, rejCount, "REJECT",dateStr)
        rejChart.rejSections = rejSections
        #Rank is NOT guaranteed to be consistent with their original placement
        log.info("Writing albums to reject chart in order by date")
        for date in datelist:
            for album in dates[date]:
                rejChart.addAlbum(album.idNo, i)
                i += 1
        return rejChart
    else:
        print("ERROR: COULD NOT FIND CHART")
def getTimestamp(albumLink):
    log.debug("Requesting {l}".format(l=albumLink))
    albPage = requests.get(albumLink)
    log.debug("Status code: {c}".format(c=albPage.status_code))
    albInfo = html.fromstring(albPage.content)
    trList = albInfo.xpath("/html/body/div[2]/div[2]/div/div[2]/table/tr/td[2]/p[1]/text()")
    durationRE = re.compile(r"total(?:[\ ]|[\ ]\w+[\ ])time.*", re.IGNORECASE)
    temp = []
    timestampRE =re.compile(r"\d{2,3}\:\d{2}")
    for j in trList:
        temp = temp + re.findall(durationRE, j)
    if len(temp) == 1:
        log.debug("Found {t} timestamp, using it".format(t=len(temp)))
        timestamp = re.search(timestampRE, temp[0])
        if timestamp:
            timestamp = timestamp.group(0)
    elif len(temp) == 2:
        log.debug("Found {t} timestamps, adding them together".format(t=len(temp)))
        ts1 = re.search(timestampRE, temp[0]).group(0)
        ts2 = re.search(timestampRE, temp[1])
        if ts2:
            ts2 = ts2.group(0)
            timestamp = addTimes(ts1, ts2)
        else:
            timestamp = ts1
    else:
        log.debug("No clear timestamps found, manually calculating")
        #trackTimeRE = re.compile(r"(?<=\d\.).*(\d{2}\:\d{2})|(?<=\d\.).*(\d{1}\:\d{2})")
        #I miss this regex. I put a lot of time into it. But it doesn't practically work. Rip
        trackTimeRE = re.compile(r"\d{1,2}\:\d{2}")
        accTime = "0:00"
        for track in trList:
            trLen = re.findall(trackTimeRE, track)
            if len(trLen) > 0:
                for ts in trLen:
                    accTime = addTimes(accTime, ts)
        if accTime == "0:00":
            accTime = "Unknown"
        timestamp = accTime
    log.info("Found timestamp: {t}".format(t=timestamp))
    return timestamp
def readChart(chartsheet, chartname, listeners, overwrite=False):
    log.info("Reading chart {name}.".format(name=chartname))
    #chartbook = load_workbook(filename)
    #chartsheet = chartbook[chartname]
    i = 1
    max_row=len([row for row in chartsheet if not all([cell.value is None for cell in row])]) 
    with tqdm.tqdm(total=max_row, position=1, bar_format='{desc}', desc='readChart: Beginning process') as desc:
        for row in tqdm.tqdm(chartsheet.iter_rows(),total=max_row, unit="album", position=0):#for every row
            if i == 1:
                i += 1
                pass
            else:
                if i > max_row:
                    break
                log.info("Reading row {rownum}".format(rownum=i))
                artist = str(row[0].value)
                artistLink = str(row[0].hyperlink.target)
                albumLink = str(row[1].hyperlink.target)
                albumName = str(row[1].value)
                year = str(row[2].value)
                genre = str(row[3].value)
                rating = str(row[4].value)
                noRatings = str(row[5].value)
                duration = str(row[6].value)
                country =  str(row[7].value)
                log.debug("Found {artist} - {album}".format(artist=artist, album=albumName))
                desc.set_description("readChart: Processing {albumName} by {artist} | Currently reading metadata".format(albumName=albumName, artist=artist))
                listenInfo = [None]*len(listeners)
                nextCol = 8
                for j in range(len(listeners)):
                    listenInfo[j] = str(row[nextCol].value)
                    nextCol = nextCol + 1
                log.debug("Found listen values {vals}".format(vals=str(listenInfo)))
                onSpotify = str(row[nextCol].value)
                if onSpotify == "Yes":
                    onSpotify = True
                    spLink = str(row[nextCol].hyperlink.target)
                else:
                    onSpotify = False
                    spLink = ""
                bought = str(row[nextCol + 1].value)
                inLibrary = str(row[nextCol + 2].value)
                houseListened = str(row[nextCol + 3].value)
                getID = re.compile(r"(?<=\?id\=).*")
                gettingID = re.search(getID, albumLink)
                albumID= gettingID.group(0)
                if albumID in albumLib.keys():
                    desc.set_description("readChart: Processing {albumName} by {artist} | Album found in library, syncing.".format(albumName=albumName, artist=artist))
                    log.info("Found albumID in library")
                    log.debug("Album ID: {albumID}".format(albumID=albumID))
                    album = albumLib[albumID]
                    album.syncListened(listenInfo.copy(), listeners)
                    album.bought = bought
                    album.links.inLibrary=inLibrary
                    album.houseListened=houseListened
                    if overwrite:
                        log.info("Overwriting spotify info")
                        if onSpotify:
                            album.links.onSpotify = True
                            album.links.spotifyLink = spLink
                        else:
                            album.links.onSpotify = False
                            album.links.spotifyLink = ""
                        log.debug("OnSpotfify: {spstatus}, link: {spLink}".format(spstatus=onSpotify, spLink = spLink))
                        log.debug("Duration: {dur}".format(dur=duration))
                        album.duration = duration
                    album = albumLib[albumID]
                    i += 1
                else:
                    #NEED NEW REGEXES
                    #desc.set_description("readChart: Processing %s by %s | Album not in library, fetching info" % (artist, albumName))
                    desc.set_description("readChart: Processing {albumName} by {artist} | Album not in library, fetching info.".format(albumName=albumName, artist=artist))
                    log.info("albumID not found in library")
                    log.debug("Album ID: {albumID}".format(albumID=albumID))
                    album = Album(albumID, albumName, artist, year)
                    album.links.albumLink = albumLink
                    album.links.artistLink = artistLink
                    log.debug("Requesting {link}".format(link=albumLink))
                    apage = requests.get(albumLink)
                    log.debug("Status code: {code}".format(code=apage.status_code))
                    acont = apage.content
                    ahtml = html.fromstring(acont)
                    rating = ahtml.xpath('//*[@id="avgRatings_1"]/text()')[0]
                    noRatings = ahtml.xpath('//*[@id="nbRatings_1"]/text()')[0]
                    #desc.set_description("readChart: Processing %s by %s | Album not in library, syncing info" % (artist, albumName))
                    desc.set_description("readChart: Processing {albumName} by {artist} | Album not in library, syncing info.".format(albumName=albumName, artist=artist))
                    album.rating = rating
                    album.noRatings = noRatings
                    album.genre = genre
                    album.setListeningInfo(listenInfo.copy(), listeners)
                    album.bought = bought
                    album.links.inLibrary=inLibrary
                    album.houseListened=houseListened
                    album.duration = duration
                    sleep(1) #Must sleep to avoid clogging the site
                    log.debug("Requesting {link}".format(link=artistLink))
                    artPage = requests.get(artistLink)
                    log.debug("Status code: {sc}".format(sc=artPage.status_code))
                    artInfo = html.fromstring(artPage.content)
                    headerInfo = artInfo.xpath("/html/body/div[2]/div[2]/div/h2/text()")
                    country = headerInfo[0].split(' • ')[1]
                    album.country = country
                    log.info("Retrieving Spotify status from API")
                    if onSpotify:
                        album.links.onSpotify = True
                        album.links.spotifyLink = spLink
                    else:
                        album.links.onSpotify = False
                        album.links.spotifyLink = ""
                    log.debug("OnSpotify: {spstatus}, link: {spLink}".format(spstatus=onSpotify, spLink = spLink))
                    log.debug("Duration: {dur}".format(dur=duration))
                    albumLib[albumID] = album
                    i += 1
    log.info("Finished after reading {numrows} rows".format(numrows=str(i)))
    return 


def updateChart(name, spObj, listeners, newLink="", updateSpotify=True, updateDuration=True):
    dateStr = date.today().strftime("%m/%d/%Y")
    if name not in chartLib.keys():
        print("Could not find chart!")
        return
    chart = chartLib[name]
    if chart.isRejectChart:
        print("you cannot update a reject chart")
        return
    if newLink == "":
        newLink = chart.link
    log.info("Updating chart {name} with the listeners {listeners} from {link}".format(name=name, listeners=str(listeners), link=newLink))
    log.debug("Requesting {link}".format(link=newLink))
    reqChart = requests.get(newLink)
    log.debug("Status code: {code}".format(code=reqChart.status_code))
    parsed = html.fromstring(reqChart.content)
    rawAlbums = parsed.xpath('/html/body/div[2]/div[2]/div[1]/table[1]/tr')
    oldAlbums = set(chart.resize(len(rawAlbums))) #this explicitly resizes the referenced old chart, returns the old albums
    i = 0
    with tqdm.tqdm(total=len(rawAlbums), position=1, bar_format='{desc}', desc='updateChart: Beginning process') as desc:
        for albumTR in tqdm.tqdm(rawAlbums, position=0, unit="album", total=len(rawAlbums)):
            i += 1
            sleep(1)
            albumLink = "http://www.progarchives.com/" + albumTR[3][0].attrib['href'] 
            artistLink = "http://www.progarchives.com/" + albumTR[3][2].attrib['href']
            getID = re.compile(r"(?<=\?id\=).*")
            gettingID = re.search(getID, albumLink)
            albumID= gettingID.group(0)
            if albumID not in albumLib.keys():
                log.info("AlbumID not found in library")
                log.debug("AlbumID: {id}".format(id=albumID))
                albumName = albumTR[3][0][0].text
                artistName = albumTR[3][2].text
                desc.set_description("updateChart: Processing {albumName} by {artist} | Currently reading metadata".format(albumName=albumName, artist=artistName))
                genre = albumTR[4][0].text
                yrRE = re.compile(r"\d{4}")
                yearDump = albumTR[4].text_content()
                year = re.search(yrRE, yearDump).group(0)
                albumRating = albumTR[2][2].text
                noRatings = albumTR[2][3].text
                #new album!!!
                desc.set_description("updateChart: Processing {albumName} by {artist} | Creating new album entry".format(albumName=albumName, artist=artistName))
                album = Album(albumID, albumName, artistName, year)
                #set already fetched info
                album.genre = genre
                album.rating = albumRating
                album.noRatings = noRatings
                album.links.artistLink = artistLink
                album.links.albumLink = albumLink
                log.info("Setting listening info for {listeners}".format(listeners=listeners))
                album.setListeningInfo(['?']*len(listeners), listeners)
                if album.links.inLibrary == False:
                    desc.set_description("updateChart: Processing {albumName} by {artist} | Searching for album on disc".format(albumName=albumName, artist=artistName))
                    album.links.findInLibrary()
                desc.set_description("updateChart: Processing {albumName} by {artist} | Finding timestamp".format(albumName=albumName, artist=artistName))
                album.duration = getTimestamp(albumLink)
                log.debug("Requesting {link}".format(link=artistLink))
                artPage = requests.get(artistLink)
                log.debug("Status code: {sc}".format(sc=artPage.status_code))
                artInfo = html.fromstring(artPage.content)
                headerInfo = artInfo.xpath("/html/body/div[2]/div[2]/div/h2/text()")
                country = headerInfo[0].split(' • ')[1]
                album.country = country
                desc.set_description("updateChart: Processing {albumName} by {artist} | Updating Spotify status".format(albumName=albumName, artist=artistName))
                album.links.findSpotify(spObj)
                albumLib[albumID] = album
                chart.addAlbum(albumID, i)
            else: #If the album has been logged before:
                log.info("Album found in library!")
                album = albumLib[albumID]
                desc.set_description("updateChart: Processing {albumName} by {artist} | Updating album metadata".format(albumName=album.title, artist=album.artist))
                prevRating = album.rating
                album.rating = albumTR[2][2].text #new rating
                log.debug("Previous rating: {pr}. Current rating: {r}".format(pr=prevRating, r=album.rating))
                prevRatingCount = album.noRatings
                log.debug("Previous # of ratings: {pr}. Current # of ratings: {r}".format(pr=prevRatingCount, r=album.noRatings))
                album.noRatings = albumTR[2][3].text #new rating count
                albumName = albumTR[3][0][0].text
                if albumName != album.title:
                    log.debug("Album name changed. Previously was {pt}, now is {t}".format(pt=album.title, t=albumName))
                    album.title=albumName
                artistName = albumTR[3][2].text
                if artistName != album.artist:
                    log.debug("Album artist changed. Previously was {pa}, now is {a}".format(pa=album.artist, a=artistName))
                    album.title=albumName
                genre = albumTR[4][0].text
                if genre != album.genre:
                    log.debug("Album genre changed. Previously was {pg}, now is {g}".format(pg=album.genre, g=genre))
                    album.genre=genre
                yrRE = re.compile(r"\d{4}")
                yearDump = albumTR[4].text_content()
                year = re.search(yrRE, yearDump).group(0)
                if year != album.year:
                    log.debug("Album year changed. Previously was {py}, now is {y}".format(py=album.year, y=year))
                    album.year=year
                if album.links.inLibrary == False:
                    desc.set_description("updateChart: Processing {albumName} by {artist} | Searching for album on disc".format(albumName=albumName, artist=artistName))
                    album.links.findInLibrary()
                if updateSpotify: #update spotify maybe
                    desc.set_description("updateChart: Processing {albumName} by {artist} | Updating Spotify status".format(albumName=albumName, artist=artistName))
                    album.links.findSpotify(spObj)
                if updateDuration:
                    desc.set_description("updateChart: Processing {albumName} by {artist} | Updating timestamp".format(albumName=albumName, artist=artistName))
                    album.duration = getTimestamp(albumLink)
                chart.addAlbum(albumID, i) #
    rejectAlbums = list(oldAlbums - set(chart.entries))
    with tqdm.tqdm(total=len(rejectAlbums), position=1, bar_format='{desc}', desc='updateChart: Processing rejects') as desc:
        for album in tqdm.tqdm(rejectAlbums, total=len(rejectAlbums), position=0, unit="album"):
            log.debug("Rejecting {title} by {artist}".format(title=albumLib[album].title, artist=albumLib[album].artist))
            desc.set_description("updateChart: Rejecting Album: {title} by {artist}".format(title=albumLib[album].title, artist=albumLib[album].artist))
            albumLib[album].addRanking(name, -1, dateStr)
    log.info("Successfully updated {chartname}".format(chartname=chart.name))
    log.info("Total albums: {acnt}, total rejected albums: {rejcnt}, total new albums: {newcnt}".format(acnt=len(chart.entries), rejcnt=len(rejectAlbums), newcnt=(len(chart.entries) - len(rejectAlbums))))
    return chart

def scanChart(name, link, spObj, listeners, updateSpotify=True, updateDuration=True):
    log.info("Scanning chart {name} from {link}".format(name=name, link=link))
    dateStr = date.today().strftime("%m/%d/%Y")
    log.debug("Sending request for {link}".format(link=link))
    reqChart = requests.get(link)
    parsed = html.fromstring(reqChart.content)
    log.debug("Status code: {code}".format(code=reqChart.status_code))
    rawAlbums = parsed.xpath('/html/body/div[2]/div[2]/div[1]/table[1]/tr')
    newChart = Chart(name, len(rawAlbums), link, dateStr)
    i = 0
    with tqdm.tqdm(total=len(rawAlbums), position=1, bar_format='{desc}', desc='Seeking album') as desc:
        for albumTR in tqdm.tqdm(rawAlbums, total=len(rawAlbums), unit="album", position=0, smoothing=0.3):
            i += 1
            sleep(1)
            albumLink = "http://www.progarchives.com/" + albumTR[3][0].attrib['href'] 
            artistLink = "http://www.progarchives.com/" + albumTR[3][2].attrib['href']
            getID = re.compile(r"(?<=\?id\=).*")
            gettingID = re.search(getID, albumLink)
            albumID= gettingID.group(0)
            albumName = albumTR[3][0][0].text
            artistName = albumTR[3][2].text
            desc.set_description("scanChart: Processing {albumName} by {artist} | Currently reading metadata".format(albumName=albumName, artist=artistName))
            genre = albumTR[4][0].text
            yrRE = re.compile(r"\d{4}")
            yearDump = albumTR[4].text_content()
            year = re.search(yrRE, yearDump).group(0)
            albumRating = albumTR[2][2].text
            noRatings = albumTR[2][3].text
            qwrRE = re.compile(r"\d\.\d*")
            qwrStr = albumTR[2][4].text
            qwr = re.search(qwrRE, qwrStr).group(0) #this should never fail
            log.debug("Found {artist} - {album}".format(artist=artistName, album=albumName))
            if albumID in albumLib.keys():
                desc.set_description("scanChart: Processing {albumName} by {artist} | Album found in library, syncing.".format(albumName=albumName, artist=artistName))
                log.info("Found albumID in library")
                log.debug("Album ID: {albumID}".format(albumID=albumID))
                album = albumLib[albumID]
                if album.links.inLibrary == False:
                    desc.set_description("scanChart: Processing {albumName} by {artist} | Searching for album on disc.".format(albumName=albumName, artist=artistName))
                    album.links.findInLibrary()
                album.rating = albumRating
                album.noRatings = noRatings
                if updateSpotify:
                    desc.set_description("scanChart: Processing {albumName} by {artist} | Updating Spotify status.".format(albumName=albumName, artist=artistName))
                    album.links.findSpotify(spObj)
                if updateDuration:
                    desc.set_description("scanChart: Processing {albumName} by {artist} | Searching for timestamp.".format(albumName=albumName, artist=artistName))
                    album.duration = getTimestamp(albumLink)
                newChart.addAlbum(albumID, i)
            else:
                desc.set_description("scanChart: Processing {albumName} by {artist} | Album not in library, fetching info.".format(albumName=albumName, artist=artistName))
                log.info("albumID not found in library")
                log.debug("Album ID: {albumID}".format(albumID=albumID))
                album = Album(albumID, albumName, artistName, year)
                album.genre = genre
                album.rating = albumRating
                album.noRatings = noRatings
                album.links.artistLink = artistLink
                album.links.albumLink = albumLink
                log.info("Setting listening info for {title}".format(title=albumName))
                album.setListeningInfo(['?']*len(listeners), listeners)
                album.qwr = qwr
                if album.links.inLibrary == False:
                    desc.set_description("scanChart: Processing {albumName} by {artist} | Searching for album on disc".format(albumName=albumName, artist=artistName))
                    album.links.findInLibrary()
                album.duration = getTimestamp(albumLink)
                log.debug("Requesting {link}".format(link=artistLink))
                artPage = requests.get(artistLink)
                log.debug("Status code: {stat}".format(stat=artPage.status_code))
                artInfo = html.fromstring(artPage.content)
                headerInfo = artInfo.xpath("/html/body/div[2]/div[2]/div/h2/text()")
                country = headerInfo[0].split(' • ')[1]
                album.country = country
                desc.set_description("scanChart: Processing {albumName} by {artist} | Updating spotify status".format(albumName=albumName, artist=artistName))
                album.links.findSpotify(spObj)
                albumLib[albumID] = album
                newChart.addAlbum(albumID, i)
        desc.set_description("scanChart: chart scan completed")
    log.info("Successfully found {nAlbs} albums for {name}".format(nAlbs=i, name=newChart.name))
    newChart.size = i
    chartLib[name] = newChart
    return newChart
def loadCharts(filename, chartLib):
    log.info("Loading chart library from {f}".format(f=filename))
    if os.path.isfile(filename):
        f = open(filename, 'rb')
        p = pickle.Unpickler(f)
        chartLib = p.load()
        f.close()
    return chartLib
def loadAlbums(filename, albumLib):
    log.info("Loading album library from {f}".format(f=filename))
    if os.path.isfile(filename):
        f = open(filename, 'rb')
        p = pickle.Unpickler(f)
        albumLib = p.load()
        f.close()
    return albumLib
def saveCharts(filename, chartLib):
    log.info("Saving chart library to {f}".format(f=filename))
    f = open(filename, 'wb')
    p = pickle.Pickler(f)
    p.dump(chartLib)
    f.close()
    return chartLib
def saveAlbums(filename, albumLib):
    log.info("Saving album library to {f}".format(f=filename))
    f = open(filename, 'wb')
    p = pickle.Pickler(f)
    p.dump(albumLib)
    f.close()
    return albumLib
def setConditionalFormatting(sheet, size, numlisteners, chart):
    log.info("Setting conditional formatting")
    #M = 9 + numlisteners -> on spotify
    #N = 10 + numlisteners -> bought
    #O = 11 + numlisteners -> in library
    #P = 12 + numlisteners -> "house listened"
    if chart.isRejectChart:
        log.debug("Reject chart found, adjusting chart size")
        size += len(chart.rejSections)
    if size == 0:
        size = 1
    f1 = Font(bold=True)
    noListenersStr = "AND($"
    for i in range(numlisteners):
        thisCol = get_column_letter(9 + i)
        noListenersStr = noListenersStr + thisCol + '2="No",$'
    noListenersStr = noListenersStr[:-2] + ")"
    noListenersRule = FormulaRule(formula=[noListenersStr], font=f1)
    f2 = Font(italic=True)
    noStreamingStr = "AND($" + get_column_letter(9 + numlisteners) + '2="No", $' + get_column_letter(11 + numlisteners) + '2="Yes")'
    noStreamingRule = FormulaRule(formula=[noStreamingStr], font=f2)
    f3 = Font(bold=True, italic=True)
    noStreamingNoListenersStr = "AND($" + get_column_letter(9 + numlisteners) + '2="No", $' + get_column_letter(11 + numlisteners) + '2="Yes", AND($'
    for i in range(numlisteners):
        thisCol = get_column_letter(9 + i)
        noStreamingNoListenersStr = noStreamingNoListenersStr + thisCol + '2="No",$'
    noStreamingNoListenersStr = noStreamingNoListenersStr[:-2] + "))"
    noStreamingNoListenersRule= FormulaRule(formula=[noStreamingNoListenersStr], font=f3)
    f4 = Font(color='FF0000')
    noStreamingNoLibraryStr = "AND($" + get_column_letter(9 + numlisteners) + '2="No", $' + get_column_letter(11 + numlisteners) + '2="No")'
    noStreamingNoLibraryRule = FormulaRule(formula=[noStreamingNoLibraryStr], font=f4)
    f5 = Font(color='FF0000', bold=True)
    noStreamingNoLibraryNoListenersStr = "AND(AND($" + get_column_letter(9 + numlisteners) + '2="No",$' + get_column_letter(11 + numlisteners) + '2="No"), AND($'
    for i in range(numlisteners):
        thisCol = get_column_letter(9 + i)
        noStreamingNoLibraryNoListenersStr = noStreamingNoLibraryNoListenersStr + thisCol + '2="No",$'
    noStreamingNoLibraryNoListenersStr = noStreamingNoLibraryNoListenersStr[:-2] + '))'
    noStreamingNoLibraryNoListenersRule= FormulaRule(formula=[noStreamingNoLibraryNoListenersStr], font=f5)
    f6 = Font(strikethrough=True)
    allListenedStr = "AND($"
    for i in range(numlisteners):
        thisCol = get_column_letter(9 + i)
        allListenedStr = allListenedStr + thisCol + '2="Yes",$'
    allListenedStr = allListenedStr[:-2] + ')'
    allListenedRule = FormulaRule(formula=[allListenedStr], font=f6)
    f7 = Font(strikethrough=True, italic=True)
    allListenedNoStreamingStr = "=AND($" + get_column_letter(9 + numlisteners) + '2="No", $' + get_column_letter(11+numlisteners) + '2="Yes", AND($'
    for i in range(numlisteners):
        thisCol = get_column_letter(9 + i)
        allListenedNoStreamingStr = allListenedNoStreamingStr + thisCol + '2="Yes",$'
    allListenedNoStreamingStr = allListenedNoStreamingStr[:-2] + "))"
    allListenedNoStreamingRule = FormulaRule(formula=[allListenedNoStreamingStr], font=f7)
    f8 = Font(strikethrough=True)
    fill = PatternFill(bgColor='D9D9D9')
    allListenedTogetherStr = "$" + get_column_letter(12 + numlisteners) + '2="Yes"'
    allListenedTogetherRule = FormulaRule(formula=[allListenedTogetherStr], font=f8, fill=fill)
    fill = PatternFill(bgColor="b7e1cd")
    dxf = DifferentialStyle(fill=fill)
    listeningBlockYesRule=Rule(type='containsText', text='Yes', dxf=dxf)
    fill1 = PatternFill(bgColor="EA9999")
    dxf = DifferentialStyle(fill=fill1)
    listeningBlockNoRule=Rule(type='containsText', text='No', dxf=dxf)
    log.info("Writing conditional formatting onto every row")
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noStreamingRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noListenersRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noStreamingNoListenersRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noStreamingNoLibraryRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noStreamingNoLibraryNoListenersRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), allListenedRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), allListenedNoStreamingRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), allListenedTogetherRule)
    sheet.conditional_formatting.add('A2:Q{end}'.format(end=size+1), noStreamingNoListenersRule)
    sheet.conditional_formatting.add('{lstart}2:{lend}{end}'.format(end=size+1, lstart=get_column_letter(9), lend=get_column_letter(9+numlisteners-1)), listeningBlockYesRule)
    sheet.conditional_formatting.add('{lstart}2:{lend}{end}'.format(end=size+1, lstart=get_column_letter(9), lend=get_column_letter(9+numlisteners-1)), listeningBlockNoRule)
    return sheet
def setColumnWidths(sheet):
    log.info("Setting column widths")
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))   
    dim_holder = DimensionHolder(worksheet=sheet)
    for col in range(sheet.min_column, sheet.max_column + 1):
        width = dims[col] * 1.23
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=width)
    sheet.column_dimensions = dim_holder
def setup(id, sec, library, albumlib, chartlib):
    with open("settings.ini", 'w') as file:
        file.write('[settings]\n')
        file.write('SPOTIPY_CLIENT_ID={id} \n'.format(id=id))
        file.write('SPOTIPY_CLIENT_SECRET={sec} \n'.format(sec=sec))
        file.write('LIBRARY_DIR={library} \n'.format(library=library))
        file.write('CHARTLIB={clib}'.format(clib=chartlib + '.pkl\n'))
        file.write('ALBUMLIB={clib}'.format(clib=albumlib + '.pkl'))
def main():
    #USE ARGPARSE
    parser = argparse.ArgumentParser("PAScraper")
    parser.add_argument('--verbose', '-b', action="store_true", help="Enables all info messages")
    parser.add_argument('--debug', '-d', action="store_true", help="Enables all debug messages")
    parser.add_argument('--log', '-l', action="store", nargs=1, help="Redirects output to given file")
    parser.add_argument('--version', '-v', action="store_true", help="Displays version")
    subparsers = parser.add_subparsers(dest='command', help='Commands to run', required=False)
    
    parser_setup = subparsers.add_parser('setup')
    parser_setup.add_argument('id', action='store', nargs=1, help="Provide a Client ID for the Spotify API")
    parser_setup.add_argument('sec', action='store', nargs=1, help="Provide a Secret for the Spotify API")
    parser_setup.add_argument('library', action='store', nargs=1, help="Provide a directory for your music library")
    parser_setup.add_argument('--albumlib', '-al', action='store', nargs=1, help='album library filename (excluding extension)')
    parser_setup.add_argument('--chartlib', '-cl', action='store', nargs=1, help='chart library filename (excluding extension)')
    
    parser_readchart = subparsers.add_parser('readchart')
    parser_readchart.add_argument('filename', nargs=1, action="store", help="Filename of workbook (.xlsx)")
    parser_readchart.add_argument('chartname', nargs=1, action="store", help="Name of chart in workbook")
    parser_readchart.add_argument('listeners', nargs='*', action="store", help="Enter names of listeners [order matters]")
    parser_readchart.add_argument('--overwrite', '-o', action="store_true", help="Overwrite Spotify status and Duration")
    
    parser_scanchart = subparsers.add_parser('scanchart')
    parser_scanchart.add_argument('chartname', nargs=1, action='store', help='Name of chart')
    parser_scanchart.add_argument('link', nargs=1, action='store', help='Progarchives chart link')
    parser_scanchart.add_argument('listeners', nargs='*', action='store', help='Listeners for chart')
    parser_scanchart.add_argument('--update-spotify', '-us', action='store_true', help='Update spotify for logged albums')
    parser_scanchart.add_argument('--update-duration', '-ud', action='store_true', help='Update duration for logged albums')
    parser_scanchart.add_argument('--write', '-w', action='store', nargs=1, help='Write scanned chart to file')
    parser_scanchart.add_argument('--new', '-n', action='store_true', help='Create new spreadsheet')
    
    parser_updatechart = subparsers.add_parser('updatechart')
    parser_updatechart.add_argument('chartname', nargs=1, action='store', help='Name of chart')
    parser_updatechart.add_argument('listeners', nargs='*', action='store',help='Listeners for chart')
    parser_updatechart.add_argument('--newlink',  nargs=1, action='store', help='Link to update chart with')
    parser_updatechart.add_argument('--update-spotify', action='store_true', help='Update spotify')
    parser_updatechart.add_argument('--update-duration', action='store_true', help='Update duration')
    parser_updatechart.add_argument('--write', '-w', action='store', nargs=1, help='Write scanned chart to file')

    parser_writechart = subparsers.add_parser('writechart')
    parser_writechart.add_argument('chartname', nargs=1, action='store', help='Name of chart')
    parser_writechart.add_argument('listeners', nargs='*', action='store',help='Listeners for chart')
    parser_writechart.add_argument('filename', nargs=1, action='store', help= 'Name of workbook')
    parser_writechart.add_argument('--new', '-n', action='store_true', help='Create new sheet')

    parser_rejchart = subparsers.add_parser('getrejectchart')
    parser_rejchart.add_argument('chartname', nargs=1, action='store', help='Name of base chart')
    parser_rejchart.add_argument('listeners', nargs='*', action='store', help='Listeners')
    parser_rejchart.add_argument('--write', '-w', nargs=1, action='store', help='File to write to (.xlsx)')
    parser_rejchart.add_argument('--new', '-n', action='store_true', help='Write to new sheet')

    parser_updateworkbook = subparsers.add_parser('updateworkbook')
    parser_updateworkbook.add_argument('filename', nargs=1, action='store', help='Filename of workbook to open (.xlsx)')
    parser_updateworkbook.add_argument('listeners', nargs='*', action='store', help='listeners for EVERY chart')
    parser_updateworkbook.add_argument('--update-spotify', action='store_true', help='Update spotify')
    parser_updateworkbook.add_argument('--update-duration', action='store_true', help='Update duration')
    parser_updateworkbook.add_argument('--write', '-w', action='store', nargs=1, help='Write scanned chart to file')
    parser_updateworkbook.add_argument('--new', '-n', action='store', nargs=1, help='Write to new file')

    parser_readworkbook = subparsers.add_parser('readworkbook')
    parser_readworkbook.add_argument('filename', nargs=1, action='store', help='Filename of workbook to open (.xlsx)')
    parser_readworkbook.add_argument('listeners', nargs='*', action='store', help='Listeners for EVERY chart')
    parser_readworkbook.add_argument('--overwrite', '-o', action='store_true', help='Overwrite stored duration and spotify status')
    
    parser_scancharts = subparsers.add_parser('scancharts')
    parser_scancharts.add_argument('chartfile', nargs=1, action='store', help='File with chart names and links separated by ::')
    parser_scancharts.add_argument('listeners', nargs='*', action='store', help='Listeners for chart')
    parser_scancharts.add_argument('--update-spotify', '-us', action='store_true', help='Update spotify for logged albums')
    parser_scancharts.add_argument('--update-duration', '-ud', action='store_true', help='Update duration for logged albums')
    parser_scancharts.add_argument('--write', '-w', action='store', nargs=1, help='Write to a given workbook')
    parser_scancharts.add_argument('--new', '-n', action='store', nargs=1,help='Save workbook as new file')
    
    a= parser.parse_args()
    if a.version:
        header = '| C H A R T  H E L P E R |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        print("Current version: 1.0") 
        print("Created by Hamish Robb, 2020-2022")

    if a.verbose:
        log.handlers[1].setLevel(logging.INFO)
    elif a.debug:
        log.handlers[1].setLevel(logging.DEBUG)
    else:
        log.handlers[1].setLevel(logging.WARNING)
    chartlibdir = ""
    albumlibdir = ""
    if a.command == 'setup':
        if not a.albumlib:
            albumlib = 'albumLib'
        else:
            albumlib = a.albumlib[0]
        if not a.chartlib:
            chartlib = 'chartLib'
        else:
            chartlib =  a.chartlib[0]
        if not os.path.isdir(a.library[0]):
            log.error("Library directory {l} could not be validated".format(l=a.library[0]))
            return -1
        setup(a.id[0], a.sec[0], a.library[0], albumlib, chartlib)
    if not os.path.isfile('settings.ini'):
        log.error('settings.ini does not exist. Try running setup')
        return -1
    with open('settings.ini', 'r') as file:
        lines = file.readlines()
        if len(lines) != 6:
            log.error("settings.ini is not well formed. Try running setup")
            return -1
        chartlibdir = lines[4].split('=')[1].strip('\n')
        albumlibdir = lines[5].split('=')[1].strip('\n')
    auth_handler = SpotifyClientCredentials(config('SPOTIPY_CLIENT_ID'), config('SPOTIPY_CLIENT_SECRET'))
    sp = spotipy.Spotify(auth_manager=auth_handler)
    wb = Workbook()
    global albumLib 
    global chartLib
    albumLib= loadAlbums(albumlibdir, albumLib)
    chartLib = loadCharts(chartlibdir, chartLib)
    
    if a.command == 'readchart':
        header = '| R E A D  C H A R T |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        if not os.path.isfile(a.filename[0]):
            log.error('{f} is not a valid file'.format(f=a.filename[0]))
            sys.exit(1)
        wb = load_workbook(a.filename[0])
        sheet = wb[a.chartname[0]]
        readChart(sheet, a.chartname[0], a.listeners, a.overwrite)
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == 'scanchart':
        header = '| S C A N  C H A R T |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        chart = scanChart(a.chartname[0], a.link[0], sp, a.listeners, a.update_spotify, a.update_duration)
        if a.write:
            if len(wb.chartsheets) == 0:
                sheet = wb.active
                sheet.title = a.chartname[0]
            else:
                if a.new:
                    sheet = wb.create_sheet(a.chartname[0])
                else:
                    sheet = wb[a.chartname[0]] 
            listeners = a.listeners
            if not a.listeners:
                listeners = ['Listened?']
            chart.writeChart(sheet, listeners)
            chartLib[chart.name] = chart
            setColumnWidths(sheet)
            sheet.freeze_panes = sheet['C2']
            sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
            wb.save(a.write[0])
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == 'updatechart':
        header = '| U P D A T E  C H A R T |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        if a.newlink:
            link = a.newlink[0]
        else:
            link = ""
        chart = updateChart(a.chartname[0], sp, a.listeners, link, a.update_spotify, a.update_duration)
        if a.write:
            if os.path.isfile(a.write[0]):
                wb = load_workbook(a.write[0])
            if len(wb.chartsheets) == 0:
                sheet = wb.active
                sheet.title = a.chartname[0]
            else:
                if a.new:
                    sheet = wb.create_sheet(a.chartname[0])
                else:
                    sheet = wb[a.chartname[0]] 
            listeners = a.listeners
            if not a.listeners:
                listeners = ['Listened?']
            chart.writeChart(sheet, listeners)
            setColumnWidths(sheet)
            sheet.freeze_panes = sheet['C2']
            sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
            wb.save(a.write[0])
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == 'writechart':
        header = '| W R I T E  C H A R T |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        chart = chartLib[a.chartname[0]]
        if os.path.isfile(a.filename[0]):
            wb = load_workbook(a.filename[0])
            if len(wb.worksheets) == 0:
                sheet = wb.active
                sheet.title = a.chartname[0]
            else:
                if a.new:
                    sheet = wb.create_sheet(a.chartname[0])
                else:
                    sheet = wb[a.chartname[0]] 
            listeners = a.listeners
            if not a.listeners:
                listeners = ['Listened?']
            chart.writeChart(sheet, listeners)
            setColumnWidths(sheet)
            sheet.freeze_panes = sheet['C2']
            sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
            wb.save(a.filename[0])
    elif a.command == 'getrejectchart':
        header = '| G E T  R E J E C T  C H A R T |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        chart = getRejectChart(a.chartname[0])
        if os.path.isfile(a.write[0]):
            wb = load_workbook(a.write[0])
            if len(wb.worksheets) == 0:
                sheet = wb.active
                sheet.title = a.chartname[0]
            else:
                if a.new:
                    sheet = wb.create_sheet(a.chartname[0])
                else:
                    sheet = wb[a.chartname[0]] 
            listeners = a.listeners
            if not a.listeners:
                listeners = ['Listened?']
            chart.writeChart(sheet, listeners)
            setColumnWidths(sheet)
            sheet.freeze_panes = sheet['C2']
            sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
            wb.save(a.write[0])
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == "updateworkbook":
        header = '| U P D A T E  W O R K B O O K |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        wb = load_workbook(a.filename[0])
        listeners = a.listeners
        if not a.listeners:
            listeners = ['Listened?']
        for sheet in wb.worksheets:
            if sheet.title in chartLib.keys():
                if chartLib[sheet.title].isRejectChart:
                    pass
            else:
                log.error("Sheet {title} was not found in chartlib".format(title=sheet.title))
            chart = updateChart(sheet.title, sp, listeners, updateSpotify=a.update_spotify, updateDuration=a.update_duration)
            if a.write:
                chart.writeChart(sheet,listeners)
                sheet.freeze_panes = sheet['C2']
                sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
                setColumnWidths(sheet)
        if a.new:
            wb.save(a.new[0])
        else:
            wb.save(a.filename[0])
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == "readworkbook":
        wb = load_workbook(a.filename[0])
        listeners = a.listeners
        if not a.listeners:
            listeners = ['Listened?']
        for sheet in wb.worksheets:
            readChart(sheet, sheet.title, listeners, a.overwrite)
        saveAlbums(albumlibdir, albumLib)
        saveCharts(chartlibdir, chartLib)
    elif a.command == 'scancharts':
        header = '| S C A N  C H A R T S |'
        rems = (termsize - len(header)) // 3
        headerPad = "~_~"*(rems//2)
        header = headerPad + header + headerPad
        print(header)
        listeners = a.listeners
        if not a.listeners:
            listeners = ['Listened?']
        if a.write:
            if os.path.isfile(a.write[0]):
                wb = load_workbook(a.write[0])
        with open(a.chartfile[0], 'r') as chartfile:
            lines = chartfile.readlines()
            for line in lines:
                parsedline = line.split('::')
                name = parsedline[0]
                link = parsedline[1]
                print("Now scanning: {s}".format(s=name))
                chart = scanChart(name, link, sp, listeners, a.update_spotify, a.update_duration)
                if a.write:
                    if name not in wb.sheetnames:
                        sheet = wb.create_sheet(name)
                    else:
                        sheet = wb[name]
                    chart.writeChart(sheet, listeners)
                    sheet.freeze_panes = sheet['C2']
                    sheet = setConditionalFormatting(sheet, chart.size, len(listeners), chart)
                    setColumnWidths(sheet)
                saveAlbums(albumlibdir, albumLib)
                saveCharts(chartlibdir, chartLib)
                if a.write:
                    if a.new:
                        wb.save(a.new[0])
                    else:
                        wb.save(a.write[0])
    sys.exit(0)

if __name__ == "__main__":
    main()